import matplotlib.pyplot as plotter
from openpyxl import load_workbook

workbook = load_workbook(filename="sheet.xlsx")
sheet = workbook.active
A = sheet["A"]
B = sheet["B"]
X,Y=[],[]

for i in range(0,sheet.max_row):
    X.append( A[i].value)
    Y.append( B[i].value)
print(X)
print(Y)

plotter.plot(X,Y)
plotter.show()
